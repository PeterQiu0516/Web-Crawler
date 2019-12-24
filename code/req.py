from selenium import webdriver
#from openpyxl import Workbook
from openpyxl import load_workbook
import time

wb = load_workbook('1219更新基金产品表.xlsx')
ws = wb['info']
row = 2

'''
a = ws['C4']
a.value = 'Changed'
if a.value == '其他':
    print('Done')
print(a.value)
a.value = 'Changed'
'''

browser = webdriver.Chrome()

for company_j in range(1, 51):
    #company
    total = 'https://www.ifund.com.hk/en/companies'
    browser.get(total)
    time.sleep(1.7)
    company = '#root > div > div.Us48g > div > div > div > div:nth-child({0}) > a'.format(company_j)
    tmp = browser.find_element_by_css_selector(company)
    comp_url = tmp.get_attribute('href')
    time.sleep(0.7)
    browser.get(comp_url)
    time.sleep(0.7)

    #I agree
    if company_j == 1:
        butt=browser.find_element_by_css_selector('body > div:nth-child(30) > div > div > div > footer > div:nth-child(2) > button')
        butt.click()
        time.sleep(0.7)

    #funds
    time.sleep(1.7)
    n = browser.find_element_by_css_selector('#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > header > span')
    num = int(n.text)

    g = browser.find_element_by_css_selector('#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._3XPMv > span')
    page = int(g.text)

    for p in range(1, page+1):
        if p == page:
            if num%5 == 0:
                   for j in range(1, 6):
                    tt = '#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div.umSYB > div > main > div > div._3Y5kg > div:nth-child({0}) > div > a'.format(j)
                    ttt = browser.find_element_by_css_selector(tt)
                    ttt_url = ttt.get_attribute('href')
                    ID = ttt_url[-6:]
                    a = ws['A{0}'.format(row)]
                    a.value = ID
                    row = row + 1
            else:
                for j in range(1, num%5+1):
                    tt = '#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div.umSYB > div > main > div > div._3Y5kg > div:nth-child({0}) > div > a'.format(j)
                    ttt = browser.find_element_by_css_selector(tt)
                    ttt_url = ttt.get_attribute('href')
                    ID = ttt_url[-6:]
                    a = ws['A{0}'.format(row)]
                    a.value = ID
                    row = row + 1
        else:
            for j in range(1, 6):
                tt = '#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div.umSYB > div > main > div > div._3Y5kg > div:nth-child({0}) > div > a'.format(j)
                ttt = browser.find_element_by_css_selector(tt)
                ttt_url = ttt.get_attribute('href')
                ID = ttt_url[-6:]
                a = ws['A{0}'.format(row)]
                a.value = ID
                row = row + 1
            if ((page < 6)|(p < 6)):
                butt_j = browser.find_element_by_css_selector('#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._1zKlW > button:nth-child({0})'.format(p+1))
            else:
                if p == page-1:
                    butt_j = browser.find_element_by_css_selector('#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._1zKlW > button:nth-child({0})'.format(7))
                else:
                    butt_j = browser.find_element_by_css_selector('#root > div > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._1zKlW > button:nth-child({0})'.format(6))
            time.sleep(0.6)
            butt_j.click()
            time.sleep(1.6)

wb.save('1219更新基金产品表.xlsx')



