# -*- coding: utf-8 -*-
''' 
Webcrawler - A web crawling project based on selenium and openpyxl. This project gives a basic example of how to utilize webdriver to crawls fund IDs from a funding company's website: www.ifund.com.hk, you could crawl any data you want from any webpage following the similar pattern of the usage of webdriver.

Author:         Changyuan Qiu 
Contact:        peterqiu@umich.edu
Latest Update:  Nov. 12, 2020

Build:      Apart from selenium and openpyxl, you need to install chrome driver and add it to the PATH from https://sites.google.com/a/chromium.org/chromedriver/downloads for executing this script.
'''

from selenium import webdriver
from openpyxl import Workbook
# from openpyxl import load_workbook
import time


wb = Workbook()
ws = wb.active

# read from existing workbook
# wb = load_workbook('test.xlsx')
# ws = wb['info']

row = 1  # start from row 1

# If some of the websites load very slowly, try increase the sleeping time between each loading
COMPANY_LOAD_TIME = 5
BUTTON_LOAD_TIME = 2

browser = webdriver.Chrome(r"chromedriver.exe",)
# preload the company page
companyURL = r'https://www.ifund.com.hk/en/companies'
browser.get(companyURL)
time.sleep(COMPANY_LOAD_TIME)
# click "I Agree"
tmpButton = browser.find_element_by_css_selector(
    "body > div:nth-child(29) > div > div > div > footer > div:nth-child(2) > button")
tmpButton.click()
time.sleep(BUTTON_LOAD_TIME)

for companyIndex in range(1, 47): # 46 funds in total
    # company page
    # companyURL = r'https://www.ifund.com.hk/en/companies'
    browser.get(companyURL)
    time.sleep(3)
    company = '#root > div > main > div.Us48g > div > div > div > div:nth-child({0}) > a'.format(
        companyIndex)
    tmp = browser.find_element_by_css_selector(company)
    compURL = tmp.get_attribute('href')  # read company link
    browser.get(compURL)
    time.sleep(COMPANY_LOAD_TIME) # MODIFIES 

    # # I agree
    # if companyIndex == 1:
    #     butt = browser.find_element_by_css_selector(
    #         'body > div:nth-child(30) > div > div > div > footer > div:nth-child(2) > button')
    #     butt.click()
    #     time.sleep(0.7)

    # funds in a company
    numFundSelector = browser.find_element_by_css_selector(
        '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > header > span')
    numFund = int(numFundSelector.text)  # total number of funds

    # total pages for a fund
    pageSelector = browser.find_element_by_css_selector(
        '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._3XPMv > span')
    page = int(pageSelector.text)  # pages for a company

    # iterate through each page
    for p in range(1, page+1):
        if p == page:  # last page
            if numFund % 5 == 0:  # grab the ID of 5 funds, at most 5 funds on a page
                for fundIndex in range(1, 6):
                    fundSelectorString = '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div.umSYB > div > main > div > div._3Y5kg > div:nth-child({0}) > div > a'.format(
                        fundIndex)
                    fundSelector = browser.find_element_by_css_selector(
                        fundSelectorString)
                    fundURL = fundSelector.get_attribute('href')
                    # fund ID, which is the final 6 digits of their url
                    ID = fundURL[-6:]
                    a = ws['A{0}'.format(row)] # locate the cell in the table
                    a.value = ID
                    row = row + 1
            else:
                # last page, could be less than 5 funds
                for fundIndex in range(1, numFund % 5+1):
                    fundSelectorString = '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div.umSYB > div > main > div > div._3Y5kg > div:nth-child({0}) > div > a'.format(
                        fundIndex)
                    fundSelector = browser.find_element_by_css_selector(
                        fundSelectorString)
                    fundURL = fundSelector.get_attribute('href')
                    ID = fundURL[-6:]
                    a = ws['A{0}'.format(row)]
                    a.value = ID
                    row = row + 1
        else:  # click button to switch between different pages
            for fundIndex in range(1, 6):  # grab the ID of 5 funds
                fundSelectorString = '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div.umSYB > div > main > div > div._3Y5kg > div:nth-child({0}) > div > a'.format(
                    fundIndex)
                fundSelector = browser.find_element_by_css_selector(
                    fundSelectorString)
                fundURL = fundSelector.get_attribute('href')
                ID = fundURL[-6:]
                a = ws['A{0}'.format(row)]
                a.value = ID
                row = row + 1
            if ((page < 6) | (p < 6)):  # click next button
                button = browser.find_element_by_css_selector(
                    '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._1zKlW > button:nth-child({0})'.format(p+1))
            else:  # when there are more then 6 buttons, it will display only 7 buttons on a page
                if p == page-1:  # the page before the last page, click the last button
                    button = browser.find_element_by_css_selector(
                        '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._1zKlW > button:nth-child({0})'.
                        format(7))
                else:
                    button = browser.find_element_by_css_selector(
                        '#root > div > main > div.Us48g._1Nk4r > div._3lkcG.J5yPI > div > div._3qpTu > div._1zKlW > button:nth-child({0})'.format(6))
            button.click()
            time.sleep(BUTTON_LOAD_TIME)

browser.quit()
wb.save('fundID.xlsx')
